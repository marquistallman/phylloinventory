"""Importador del catalogo Excel de Colsubsidio a PostgreSQL.

Lee BODEGAS Y STOCK.xlsx (9 hojas) e inserta:
  - 48 bodegas
  - ~1,407 productos con stock, unidad y codigo de articulo

Uso:
  python -m src.importer [--excel PATH] [--dsn DSN]
"""
from __future__ import annotations

import argparse
import os
import sys
from typing import Any

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DEFAULT_EXCEL = os.path.join(os.path.dirname(__file__), "..", "BODEGAS Y STOCK.xlsx")
DEFAULT_DSN = os.getenv("DATABASE_URL", "host=localhost port=5432 dbname=inventario user=cactus password=cactus")

# Mapeo: nombre de hoja -> nombre de bodega
HOJA_A_BODEGA: dict[str, str] = {
    "STOCK ALMACEN  SUMINISTROS":    "almacen suministros",
    "STOCK ALMACEN AYB":              "almacen ayb",
    "STOCK RESTAURANTE FUENTES AYB":  "restaurante fuentes ayb",
    "STOCK RESTAURANTE FUENTES SUMIN":"restaurante fuentes suministros",
    "STOCK KIOSCO TAQUILLA AYB":      "kiosco taquilla ayb",
    "STOCK KIOSCO PISCIGIROS AYB":    "kiosco piscigiros ayb",
    "ZOOLOGICO":                      "zoologico",
    "ZOOLOGICO SUMINISTROS":          "zoologico suministros",
}


def _parse_articulo(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s == "0":
        return None
    return s


def normalizar_unidad(raw: str) -> str:
    u = (raw or "").strip().capitalize()
    if u in ("Unidad", "Kilogram", "Liter"):
        return u
    umap = {
        "kg": "Kilogram", "kilo": "Kilogram", "kilos": "Kilogram",
        "kilogramo": "Kilogram", "kilogramos": "Kilogram",
        "g": "Kilogram", "gr": "Kilogram", "gramo": "Kilogram", "gramos": "Kilogram",
        "l": "Liter", "lt": "Liter", "lts": "Liter",
        "litro": "Liter", "litros": "Liter",
        "ml": "Liter", "mililitro": "Liter", "mililitros": "Liter",
        "un": "Unidad", "und": "Unidad", "unds": "Unidad",
        "unidad": "Unidad", "unidades": "Unidad",
        "pza": "Unidad", "pzas": "Unidad", "pieza": "Unidad", "piezas": "Unidad",
        "caja": "Unidad", "cajas": "Unidad",
        "paquete": "Unidad", "paquetes": "Unidad",
        "sobre": "Unidad", "sobres": "Unidad",
        "frasco": "Unidad", "frascos": "Unidad",
        "rollo": "Unidad", "rollos": "Unidad",
    }
    return umap.get(u.lower(), "Unidad")


def importar(excel_path: str, dsn: str) -> dict[str, int]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    conn = psycopg2.connect(dsn)
    stats: dict[str, int] = {"bodegas": 0, "productos": 0, "errores": 0}

    try:
        with conn.cursor() as cur:
            # 0. Verificar que las tablas existen (init.sql se haya corrido)
            cur.execute("""
                SELECT table_name FROM information_schema.tables
                WHERE table_schema = 'public'
                  AND table_name IN ('bodegas', 'productos', 'stock', 'unidades')
            """)
            tables = {r[0] for r in cur.fetchall()}
            missing = {"bodegas", "productos", "stock", "unidades"} - tables
            if missing:
                raise RuntimeError(
                    f"Tablas faltantes en DB: {sorted(missing)}. "
                    f"Ejecuta db/init.sql contra la base de datos "
                    f"(ej: docker compose down -v && docker compose up, "
                    f"o: docker exec -i cactus_postgres psql -U cactus -d inventario < db/init.sql)."
                )

            # 1. Importar bodegas desde hoja "BODEGAS DISPONIBLES"
            ws_bodegas = wb["BODEGAS DISPONIBLES"]
            bodegas_vistas: set[str] = set()
            bodega_id_map: dict[str, int] = {}

            # Bodega default ya existe (id=1), la mapeamos
            bodega_id_map["bodega_default"] = 1

            for row in ws_bodegas.iter_rows(min_row=3, values_only=True):
                if not row or len(row) < 3:
                    continue
                nombre = str(row[2]).strip().lower() if row[2] else ""
                if not nombre or nombre in bodegas_vistas:
                    continue
                try:
                    cur.execute(
                        "INSERT INTO bodegas (nombre) VALUES (%s) ON CONFLICT (nombre) DO UPDATE SET nombre=EXCLUDED.nombre RETURNING id",
                        (nombre,),
                    )
                    bid = cur.fetchone()[0]
                    bodega_id_map[nombre] = bid
                    bodegas_vistas.add(nombre)
                    stats["bodegas"] += 1
                except Exception as e:
                    stats["errores"] += 1
                    print(f"  [ERROR] bodega '{nombre}': {e}")

            # Tambien registrar bodegas del mapeo HOJA_A_BODEGA si no existen
            for bname in HOJA_A_BODEGA.values():
                if bname in bodegas_vistas:
                    continue
                try:
                    cur.execute(
                        "INSERT INTO bodegas (nombre) VALUES (%s) ON CONFLICT (nombre) DO UPDATE SET nombre=EXCLUDED.nombre RETURNING id",
                        (bname,),
                    )
                    bid = cur.fetchone()[0]
                    bodega_id_map[bname] = bid
                    bodegas_vistas.add(bname)
                    stats["bodegas"] += 1
                except Exception as e:
                    stats["errores"] += 1
                    print(f"  [ERROR] bodega '{bname}': {e}")

            # Cache de unidades (lookup por nombre -> id)
            cur.execute("SELECT id, nombre FROM unidades")
            unidad_id_map: dict[str, int] = {r[1]: r[0] for r in cur.fetchall()}

            conn.commit()
            print(f"  Bodegas importadas: {stats['bodegas']}")

            # 2. Importar productos + stock desde las 8 hojas
            for sheet_name, bodega_nombre in HOJA_A_BODEGA.items():
                if sheet_name not in wb.sheetnames:
                    print(f"  [AVISO] Hoja '{sheet_name}' no encontrada, saltando")
                    continue

                bodega_id = bodega_id_map.get(bodega_nombre)
                if bodega_id is None:
                    print(f"  [AVISO] Bodega '{bodega_nombre}' no encontrada en DB, saltando {sheet_name}")
                    continue

                ws = wb[sheet_name]
                # Acumulamos: upsert en productos (catalogo) y luego en stock
                # (per-bodega). Lo hacemos en dos pasadas para evitar
                # conflictos con UNIQUE en (nombre) al re-correr.
                productos_rows: list[tuple] = []
                stock_rows: list[tuple] = []

                for row in ws.iter_rows(min_row=3, values_only=True):
                    if not row or len(row) < 5:
                        continue
                    # Col A=CANTIDAD(row counter), B=Nr.Articulo, C=Articulo, D=Unidad, E=SD
                    codigo = _parse_articulo(row[1]) if len(row) > 1 else None
                    nombre = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    unidad_nombre = normalizar_unidad(str(row[3])) if len(row) > 3 and row[3] else "Unidad"
                    try:
                        sd = float(row[4]) if len(row) > 4 and row[4] is not None else 0.0
                    except (ValueError, TypeError):
                        sd = 0.0

                    if not nombre:
                        continue

                    unidad_id = unidad_id_map.get(unidad_nombre)
                    if unidad_id is None:
                        cur.execute(
                            "INSERT INTO unidades (nombre) VALUES (%s) ON CONFLICT (nombre) DO UPDATE SET nombre=EXCLUDED.nombre RETURNING id",
                            (unidad_nombre,),
                        )
                        unidad_id = cur.fetchone()[0]
                        unidad_id_map[unidad_nombre] = unidad_id

                    productos_rows.append((nombre, codigo, unidad_id))
                    stock_rows.append((nombre, codigo, unidad_id, bodega_id, sd))

                # 2a. Upsert en catalogo productos
                if productos_rows:
                    execute_values(
                        cur,
                        """
                        INSERT INTO productos (nombre, codigo_articulo, unidad_id)
                        VALUES %s
                        ON CONFLICT (nombre) DO UPDATE SET
                            codigo_articulo = EXCLUDED.codigo_articulo,
                            unidad_id       = EXCLUDED.unidad_id
                        """,
                        productos_rows,
                        template="(%s, %s, %s)",
                    )

                # 2b. Upsert en stock (per-bodega)
                if stock_rows:
                    execute_values(
                        cur,
                        """
                        INSERT INTO stock (producto_id, bodega_id, stock_actual, media_kalman, varianza_kalman)
                        SELECT p.id, s.bodega_id, s.sd, s.sd, 100.0
                        FROM (VALUES %s) AS s(nombre, codigo, unidad_id, bodega_id, sd)
                        JOIN productos p ON p.nombre = s.nombre
                        ON CONFLICT (producto_id, bodega_id) DO UPDATE SET
                            stock_actual   = EXCLUDED.stock_actual,
                            media_kalman   = EXCLUDED.media_kalman,
                            actualizado_en = NOW()
                        """,
                        stock_rows,
                        template="(%s, %s, %s, %s, %s)",
                    )
                    stats["productos"] += len(stock_rows)

            conn.commit()

    finally:
        conn.close()

    return stats


def main():
    parser = argparse.ArgumentParser(description="Importar catalogo Excel a PostgreSQL")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Ruta al archivo Excel")
    parser.add_argument("--dsn", default=DEFAULT_DSN, help="DSN de PostgreSQL")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"ERROR: No se encuentra el archivo Excel: {args.excel}")
        sys.exit(1)

    print(f"Importando catalogo desde: {args.excel}")
    stats = importar(args.excel, args.dsn)
    print(f"\nResultado: {stats['bodegas']} bodegas, {stats['productos']} productos, {stats['errores']} errores")


if __name__ == "__main__":
    main()